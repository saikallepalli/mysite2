import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, DataBarRule

data_long = pd.read_csv('data_long.csv')
visits_wide = pd.read_csv('visits_wide.csv')
unique_items = pd.read_csv('unique_items.csv')

N_VISITS = len(visits_wide)
N_DATA = len(data_long)
VROWS = N_VISITS + 1   # last row with data on Visits sheet
DROWS = N_DATA + 1     # last row with data on Data sheet

NAVY = "1F3864"
GOLD = "B8860B"
GREEN = "2E7D32"
AMBER = "C77700"
GREY = "8A8A8A"
LIGHTGREY = "F2F2F2"
WHITE = "FFFFFF"

FONT_NAME = "Arial"

def set_font(cell, size=11, bold=False, color="000000", italic=False):
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color, italic=italic)

def fill(cell, color):
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

thin = Side(style="thin", color="D9D9D9")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# =========================================================
# SHEET 1: Dashboard
# =========================================================
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False

managers = sorted(data_long['Manager'].unique().tolist())
centres = sorted(data_long['Centre'].unique().tolist())

# Title block
ws.merge_cells('A1:H2')
ws['A1'] = "Unannounced Monthly Centre Visit — Interactive Dashboard"
set_font(ws['A1'], size=18, bold=True, color=WHITE)
ws['A1'].alignment = Alignment(vertical='center', horizontal='left', indent=1)
for r in range(1,3):
    for c in range(1,9):
        fill(ws.cell(row=r, column=c), NAVY)
ws.row_dimensions[1].height = 18
ws.row_dimensions[2].height = 18

ws.merge_cells('A3:H3')
date_min = pd.to_datetime(visits_wide['Date']).min().strftime('%-d %b %Y') if len(visits_wide) else ''
date_max = pd.to_datetime(visits_wide['Date']).max().strftime('%-d %b %Y') if len(visits_wide) else ''
ws['A3'] = f"Second Level Manager field audits · {date_min}–{date_max} · Select a Manager and/or Centre below to filter every table and chart on this sheet"
set_font(ws['A3'], size=10, italic=True, color="595959")
ws.row_dimensions[3].height = 16

# Filter cells
ws['A5'] = "Second Level Manager:"
set_font(ws['A5'], bold=True, size=11)
ws['B5'] = "All Managers"
set_font(ws['B5'], bold=True, size=11, color=NAVY)
fill(ws['B5'], "FFF2CC")
ws['B5'].border = border_all
ws['B5'].alignment = Alignment(horizontal='center')

ws['D5'] = "Centre:"
set_font(ws['D5'], bold=True, size=11)
ws['E5'] = "All Centres"
set_font(ws['E5'], bold=True, size=11, color=NAVY)
fill(ws['E5'], "FFF2CC")
ws['E5'].border = border_all
ws['E5'].alignment = Alignment(horizontal='center')

ws['A6'] = "The Centre list narrows to the selected Manager's centres — if you switch Manager, reopen the Centre dropdown and pick again."
set_font(ws['A6'], size=8, italic=True, color="8A8A8A")
ws.merge_cells('A6:H6')

# Data validation lists (use a hidden Lists sheet to avoid Excel's 255-char inline-list limit)
# Also builds one centre-list column PER MANAGER so the Centre dropdown can cascade off the
# Manager selection via INDIRECT() + a named range per manager (standard Excel dependent-dropdown pattern).
ws_lists = wb.create_sheet("Lists")
ws_lists.sheet_state = "hidden"
ws_lists['A1'] = "All Managers"
for i, m in enumerate(managers, start=2):
    ws_lists.cell(row=i, column=1, value=m)
ws_lists['B1'] = "All Centres"
for i, c in enumerate(centres, start=2):
    ws_lists.cell(row=i, column=2, value=c)

import re as _re
from openpyxl.utils import get_column_letter as _gcl
from openpyxl.workbook.defined_name import DefinedName

def _sanitize(name):
    # Named ranges can't contain spaces or most punctuation; must not start with a digit.
    cleaned = _re.sub(r'[^0-9A-Za-z_]', '', name)
    if cleaned and cleaned[0].isdigit():
        cleaned = '_' + cleaned
    return cleaned or 'Mgr'

manager_centres_map = {m: sorted(data_long.loc[data_long['Manager'] == m, 'Centre'].unique().tolist()) for m in managers}

# "All Managers" -> full centre list, reusing column B so INDIRECT("Ctr_"&sanitized) resolves for every case.
wb.defined_names['Ctr_AllManagers'] = DefinedName('Ctr_AllManagers', attr_text=f"Lists!$B$1:$B${len(centres)+1}")

next_col = 3  # column C onward, one per manager
for m in managers:
    col_letter = _gcl(next_col)
    ws_lists.cell(row=1, column=next_col, value="All Centres")
    m_centres = manager_centres_map[m]
    for i, c in enumerate(m_centres, start=2):
        ws_lists.cell(row=i, column=next_col, value=c)
    range_name = f"Ctr_{_sanitize(m)}"
    wb.defined_names[range_name] = DefinedName(range_name, attr_text=f"Lists!${col_letter}$1:${col_letter}${len(m_centres)+1}")
    next_col += 1

dv_mgr = DataValidation(type="list", formula1=f"=Lists!$A$1:$A${len(managers)+1}", allow_blank=False)
dv_ctr = DataValidation(type="list", formula1='=INDIRECT("Ctr_"&SUBSTITUTE($B$5," ",""))', allow_blank=False)
ws.add_data_validation(dv_mgr)
ws.add_data_validation(dv_ctr)
dv_mgr.add(ws['B5'])
dv_ctr.add(ws['E5'])

# Internal criteria helper cells (wildcard trick for "All")
ws['A7'] = "internal:"
set_font(ws['A7'], size=8, color="BFBFBF")
ws['B7'] = '=IF($B$5="All Managers","*",$B$5)'
ws['C7'] = '=IF($E$5="All Centres","*",$E$5)'
set_font(ws['B7'], size=8, color="BFBFBF")
set_font(ws['C7'], size=8, color="BFBFBF")
ws.row_dimensions[7].hidden = True

CRIT_M = "$B$7"
CRIT_C = "$C$7"

# ---- Summary Cards (row 9-13) ----
card_specs = [
    ("Visits in View", f'=COUNTIFS(Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})', "visits matching current filter"),
    ("Overall % Need Improvement",
     f'=IFERROR((SUMIFS(Visits!$F$2:$F${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})+SUMIFS(Visits!$I$2:$I${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})+SUMIFS(Visits!$L$2:$L${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})+SUMIFS(Visits!$O$2:$O${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C}))/'
     f'((SUMIFS(Visits!$E$2:$E${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})+SUMIFS(Visits!$F$2:$F${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C}))+'
     f'(SUMIFS(Visits!$H$2:$H${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})+SUMIFS(Visits!$I$2:$I${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C}))+'
     f'(SUMIFS(Visits!$K$2:$K${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})+SUMIFS(Visits!$L$2:$L${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C}))+'
     f'(SUMIFS(Visits!$N$2:$N${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})+SUMIFS(Visits!$O$2:$O${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C}))),0)',
     '="Of "&SUM($B$18:$D$21)&" checklist items: "&SUM($C$18:$C$21)&" Need Improvement, "&SUM($D$18:$D$21)&" Not Applicable."'),
    ("Weakest Section", '=INDEX($A$18:$A$21,MATCH(MAX($E$18:$E$21),$E$18:$E$21,0))',
     '="Of "&(INDEX($B$18:$B$21,MATCH(MAX($E$18:$E$21),$E$18:$E$21,0))+INDEX($C$18:$C$21,MATCH(MAX($E$18:$E$21),$E$18:$E$21,0))+INDEX($D$18:$D$21,MATCH(MAX($E$18:$E$21),$E$18:$E$21,0)))&" items in this section: "&INDEX($C$18:$C$21,MATCH(MAX($E$18:$E$21),$E$18:$E$21,0))&" Need Improvement, "&INDEX($D$18:$D$21,MATCH(MAX($E$18:$E$21),$E$18:$E$21,0))&" Not Applicable ("&TEXT(MAX($E$18:$E$21),"0.0%")&")."'),
    ("Top Gap Item", '=B26', "most-flagged single checklist item in view"),
]

col_starts = [1,3,5,7]
ws.row_dimensions[9].height = 14
ws.row_dimensions[12].height = 28
for (label, formula, sub), cs in zip(card_specs, col_starts):
    ce = cs+1
    ws.merge_cells(start_row=9, start_column=cs, end_row=9, end_column=ce)
    c = ws.cell(row=9, column=cs, value=label)
    set_font(c, size=9, bold=True, color="595959")
    ws.merge_cells(start_row=10, start_column=cs, end_row=11, end_column=ce)
    c2 = ws.cell(row=10, column=cs, value=formula)
    set_font(c2, size=16, bold=True, color=NAVY)
    c2.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=12, start_column=cs, end_row=12, end_column=ce)
    c3 = ws.cell(row=12, column=cs, value=sub)
    set_font(c3, size=8, italic=True, color="808080")
    c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    for r in range(9,13):
        for cc in range(cs, ce+1):
            fill(ws.cell(row=r, column=cc), LIGHTGREY)
            ws.cell(row=r, column=cc).border = border_all

ws['B10'].number_format = '0'
ws['C10'].number_format = '0'
ws['C10'].number_format = '0.0%'
# fix: card 2 value cell is actually C10 due to merge start col=3
ws['C10'].number_format = '0.0%'

# ---- Section-wise stats table (rows 16-21) ----
ws['A16'] = "Section-Wise Compliance"
set_font(ws['A16'], size=13, bold=True, color=NAVY)
ws.merge_cells('A16:H16')

headers = ["Section","All is Well","Need Improvement","Not Applicable","% Need Improvement"]
for i,h in enumerate(headers):
    c = ws.cell(row=17, column=1+i, value=h)
    set_font(c, bold=True, color=WHITE, size=10)
    fill(c, NAVY)
    c.alignment = Alignment(horizontal='center')
    c.border = border_all

section_defs = [
    ("Data Check", "C", "D", "E"),
    ("Infra Check", "F", "G", "H"),
    ("Process Check", "I", "J", "K"),
    ("Aspirant Interaction", "L", "M", "N"),
]
# Visits sheet column letters: A Manager,B Centre,C Program,D Date,
# E DataCheck_AW,F DataCheck_NI,G DataCheck_NA,
# H InfraCheck_AW,I InfraCheck_NI,J InfraCheck_NA,
# K ProcessCheck_AW,L ProcessCheck_NI,M ProcessCheck_NA,
# N AspirantInt_AW,O AspirantInt_NI,P AspirantInt_NA,
# Q Q_Bottleneck,R Q_PlacementAction,S Q_BatchOnTimeAction
sec_cols = {
    "Data Check": ("E","F","G"),
    "Infra Check": ("H","I","J"),
    "Process Check": ("K","L","M"),
    "Aspirant Interaction": ("N","O","P"),
}
for ridx, sec in enumerate(["Data Check","Infra Check","Process Check","Aspirant Interaction"]):
    r = 18+ridx
    aw_col, ni_col, na_col = sec_cols[sec]
    ws.cell(row=r, column=1, value=sec)
    ws.cell(row=r, column=2, value=f'=SUMIFS(Visits!${aw_col}$2:${aw_col}${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})')
    ws.cell(row=r, column=3, value=f'=SUMIFS(Visits!${ni_col}$2:${ni_col}${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})')
    ws.cell(row=r, column=4, value=f'=SUMIFS(Visits!${na_col}$2:${na_col}${VROWS},Visits!$A$2:$A${VROWS},{CRIT_M},Visits!$B$2:$B${VROWS},{CRIT_C})')
    ws.cell(row=r, column=5, value=f'=IFERROR(C{r}/(B{r}+C{r}),0)')
    ws.cell(row=r, column=5).number_format = '0.0%'
    for cc in range(1,6):
        cell = ws.cell(row=r, column=cc)
        set_font(cell, size=10)
        cell.border = border_all
        cell.alignment = Alignment(horizontal='center' if cc>1 else 'left')
        if ridx % 2 == 1:
            fill(cell, LIGHTGREY)

# conditional data bar on %NI column
ws.conditional_formatting.add('E18:E21', DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                                                       color="C77700"))

# Bar chart referencing section stats
chart = BarChart()
chart.type = "bar"
chart.title = "% Need Improvement by Section"
chart.y_axis.title = None
chart.x_axis.title = None
chart.style = 10
data_ref = Reference(ws, min_col=5, min_row=17, max_row=21)
cats_ref = Reference(ws, min_col=1, min_row=18, max_row=21)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.height = 7
chart.width = 15
chart.legend = None
ws.add_chart(chart, "G16")

# ---- Need Improvement Spotlight (rows 23 onward) ----
ws['A23'] = "Need Improvement Spotlight — Top 15 Checklist Items"
set_font(ws['A23'], size=13, bold=True, color=NAVY)
ws.merge_cells('A23:H23')
ws['A24'] = "Ranked by number of “Need Improvement” flags among visits currently in view"
set_font(ws['A24'], size=9, italic=True, color="595959")
ws.merge_cells('A24:H24')

sp_headers = ["Section","Checklist Item","Need Improvement","Applicable N","% Need Improvement"]
for i,h in enumerate(sp_headers):
    c = ws.cell(row=25, column=1+i, value=h)
    set_font(c, bold=True, color=WHITE, size=10)
    fill(c, GOLD)
    c.alignment = Alignment(horizontal='center')
    c.border = border_all

N_ITEMS = len(unique_items)  # 49
for k in range(1, 16):
    r = 25+k
    tb_range = f"ItemStats!$G$2:$G${N_ITEMS+1}"
    a_range = f"ItemStats!$A$2:$A${N_ITEMS+1}"
    b_range = f"ItemStats!$B$2:$B${N_ITEMS+1}"
    c_range = f"ItemStats!$C$2:$C${N_ITEMS+1}"
    e_range = f"ItemStats!$E$2:$E${N_ITEMS+1}"
    f_range = f"ItemStats!$F$2:$F${N_ITEMS+1}"
    match_formula = f"MATCH(LARGE({tb_range},{k}),{tb_range},0)"
    ws.cell(row=r, column=1, value=f'=IFERROR(INDEX({a_range},{match_formula}),"")')
    ws.cell(row=r, column=2, value=f'=IFERROR(INDEX({b_range},{match_formula}),"")')
    ws.cell(row=r, column=3, value=f'=IFERROR(INDEX({c_range},{match_formula}),"")')
    ws.cell(row=r, column=4, value=f'=IFERROR(INDEX({e_range},{match_formula}),"")')
    ws.cell(row=r, column=5, value=f'=IFERROR(INDEX({f_range},{match_formula}),"")')
    ws.cell(row=r, column=5).number_format = '0.0%'
    for cc in range(1,6):
        cell = ws.cell(row=r, column=cc)
        set_font(cell, size=10)
        cell.border = border_all
        cell.alignment = Alignment(horizontal='center' if cc in (1,3,4,5) else 'left', wrap_text=(cc==2))
        if k % 2 == 0:
            fill(cell, LIGHTGREY)

ws.conditional_formatting.add(f'E26:E40', DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1, color="C77700"))

# Column widths
widths = {'A':22,'B':46,'C':16,'D':14,'E':18,'F':14,'G':14,'H':14}
for col,w in widths.items():
    ws.column_dimensions[col].width = w

ws.freeze_panes = "A8"

print("Dashboard sheet built")

# =========================================================
# SHEET: ItemStats (helper — hidden)
# =========================================================
ws_items = wb.create_sheet("ItemStats")
headers = ["Section","Item","NeedImprovement","AllIsWell","Applicable","PctNeedImprovement","TieBreak"]
for i,h in enumerate(headers):
    c = ws_items.cell(row=1, column=1+i, value=h)
    set_font(c, bold=True, color=WHITE, size=10)
    fill(c, NAVY)

for i, rowdata in unique_items.iterrows():
    r = i+2
    sec = rowdata['Section']; item = rowdata['Item']
    ws_items.cell(row=r, column=1, value=sec)
    ws_items.cell(row=r, column=2, value=item)
    ni_f = (f'=COUNTIFS(Data!$A$2:$A${DROWS},Dashboard!{CRIT_M},Data!$B$2:$B${DROWS},Dashboard!{CRIT_C},'
            f'Data!$E$2:$E${DROWS},$A{r},Data!$F$2:$F${DROWS},$B{r},Data!$G$2:$G${DROWS},"Need Improvement")')
    aw_f = (f'=COUNTIFS(Data!$A$2:$A${DROWS},Dashboard!{CRIT_M},Data!$B$2:$B${DROWS},Dashboard!{CRIT_C},'
            f'Data!$E$2:$E${DROWS},$A{r},Data!$F$2:$F${DROWS},$B{r},Data!$G$2:$G${DROWS},"All is Well")')
    ws_items.cell(row=r, column=3, value=ni_f)
    ws_items.cell(row=r, column=4, value=aw_f)
    ws_items.cell(row=r, column=5, value=f'=C{r}+D{r}')
    ws_items.cell(row=r, column=6, value=f'=IFERROR(C{r}/E{r},0)')
    ws_items.cell(row=r, column=6).number_format = '0.0%'
    ws_items.cell(row=r, column=7, value=f'=C{r}+(ROW()/100000)')
    for cc in range(1,8):
        set_font(ws_items.cell(row=r, column=cc), size=9)

ws_items.column_dimensions['A'].width = 22
ws_items.column_dimensions['B'].width = 60
for col in ['C','D','E','F','G']:
    ws_items.column_dimensions[col].width = 14
ws_items.sheet_state = "hidden"

print("ItemStats sheet built")

# =========================================================
# SHEET: Visits (one row per visit)
# =========================================================
ws_v = wb.create_sheet("Visits")
v_headers = ["Manager","Centre","Program","Date",
             "DataCheck_AW","DataCheck_NI","DataCheck_NA",
             "InfraCheck_AW","InfraCheck_NI","InfraCheck_NA",
             "ProcessCheck_AW","ProcessCheck_NI","ProcessCheck_NA",
             "AspirantInt_AW","AspirantInt_NI","AspirantInt_NA",
             "Q_Bottleneck","Q_PlacementAction","Q_BatchOnTimeAction",
             "MatchFlag","MatchRank"]
for i,h in enumerate(v_headers):
    c = ws_v.cell(row=1, column=1+i, value=h)
    set_font(c, bold=True, color=WHITE, size=10)
    fill(c, NAVY)

for i, rowdata in visits_wide.iterrows():
    r = i+2
    ws_v.cell(row=r, column=1, value=rowdata['Manager'])
    ws_v.cell(row=r, column=2, value=rowdata['Centre'])
    ws_v.cell(row=r, column=3, value=rowdata['Program'])
    ws_v.cell(row=r, column=4, value=rowdata['Date'])
    ws_v.cell(row=r, column=5, value=int(rowdata['DataCheck_AW']))
    ws_v.cell(row=r, column=6, value=int(rowdata['DataCheck_NI']))
    ws_v.cell(row=r, column=7, value=int(rowdata['DataCheck_NA']))
    ws_v.cell(row=r, column=8, value=int(rowdata['InfraCheck_AW']))
    ws_v.cell(row=r, column=9, value=int(rowdata['InfraCheck_NI']))
    ws_v.cell(row=r, column=10, value=int(rowdata['InfraCheck_NA']))
    ws_v.cell(row=r, column=11, value=int(rowdata['ProcessCheck_AW']))
    ws_v.cell(row=r, column=12, value=int(rowdata['ProcessCheck_NI']))
    ws_v.cell(row=r, column=13, value=int(rowdata['ProcessCheck_NA']))
    ws_v.cell(row=r, column=14, value=int(rowdata['AspirantInt_AW']))
    ws_v.cell(row=r, column=15, value=int(rowdata['AspirantInt_NI']))
    ws_v.cell(row=r, column=16, value=int(rowdata['AspirantInt_NA']))
    ws_v.cell(row=r, column=17, value=rowdata['Q_Bottleneck'])
    ws_v.cell(row=r, column=18, value=rowdata['Q_PlacementAction'])
    ws_v.cell(row=r, column=19, value=rowdata['Q_BatchOnTimeAction'])
    ws_v.cell(row=r, column=20, value=(f'=IF(AND(OR(Dashboard!$B$5="All Managers",A{r}=Dashboard!$B$5),'
                                        f'OR(Dashboard!$E$5="All Centres",B{r}=Dashboard!$E$5)),1,0)'))
    ws_v.cell(row=r, column=21, value=f'=IF(T{r}=1,COUNTIF($T$2:T{r},1),"")')
    for cc in range(1,22):
        set_font(ws_v.cell(row=r, column=cc), size=9)

n_visits = len(visits_wide)
last_row = n_visits+1
tbl = Table(displayName="tblVisits", ref=f"A1:U{last_row}")
tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws_v.add_table(tbl)

widths_v = {'A':20,'B':22,'C':20,'D':12,'Q':40,'R':40,'S':40,'T':10,'U':10}
for col,w in widths_v.items():
    ws_v.column_dimensions[col].width = w
for col in ['E','F','G','H','I','J','K','L','M','N','O','P']:
    ws_v.column_dimensions[col].width = 11

print("Visits sheet built")

# =========================================================
# SHEET: Data (long format raw — filterable)
# =========================================================
ws_d = wb.create_sheet("Data")
d_headers = ["Manager","Centre","Program","Date","Section","Item","Status"]
for i,h in enumerate(d_headers):
    c = ws_d.cell(row=1, column=1+i, value=h)
    set_font(c, bold=True, color=WHITE, size=10)
    fill(c, NAVY)

for i, rowdata in data_long.iterrows():
    r = i+2
    ws_d.cell(row=r, column=1, value=rowdata['Manager'])
    ws_d.cell(row=r, column=2, value=rowdata['Centre'])
    ws_d.cell(row=r, column=3, value=rowdata['Program'])
    ws_d.cell(row=r, column=4, value=rowdata['Date'])
    ws_d.cell(row=r, column=5, value=rowdata['Section'])
    ws_d.cell(row=r, column=6, value=rowdata['Item'])
    ws_d.cell(row=r, column=7, value=rowdata['Status'])
    for cc in range(1,8):
        set_font(ws_d.cell(row=r, column=cc), size=9)

n_data = len(data_long)
last_row_d = n_data+1
tbl_d = Table(displayName="tblData", ref=f"A1:G{last_row_d}")
tbl_d.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
ws_d.add_table(tbl_d)

ws_d.column_dimensions['A'].width = 20
ws_d.column_dimensions['B'].width = 22
ws_d.column_dimensions['C'].width = 20
ws_d.column_dimensions['D'].width = 12
ws_d.column_dimensions['E'].width = 18
ws_d.column_dimensions['F'].width = 70
ws_d.column_dimensions['G'].width = 16

print("Data sheet built")

# =========================================================
# SHEET: Reflections (filtered view of the 3 focus questions)
# =========================================================
ws_r = wb.create_sheet("Reflections")
ws_r.sheet_view.showGridLines = False
ws_r.merge_cells('A1:F2')
ws_r['A1'] = "Field Reflections — filtered by Dashboard selection"
set_font(ws_r['A1'], size=14, bold=True, color=WHITE)
for cc in range(1,7):
    fill(ws_r.cell(row=1, column=cc), NAVY)
    fill(ws_r.cell(row=2, column=cc), NAVY)
ws_r['A3'] = "This list updates automatically based on the Manager / Centre selected on the Dashboard sheet."
set_font(ws_r['A3'], size=9, italic=True, color="595959")
ws_r.merge_cells('A3:F3')

r_headers = ["Centre","Manager","Program","Date","Key bottleneck resolved","Placement % action","Batch on-time action"]
for i,h in enumerate(r_headers):
    c = ws_r.cell(row=5, column=1+i, value=h)
    set_font(c, bold=True, color=WHITE, size=10)
    fill(c, GOLD)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

for k in range(1, n_visits+1):
    r = 5+k
    rank_range = f"Visits!$U$2:$U${last_row}"
    match_f = f'MATCH({k},{rank_range},0)'
    ws_r.cell(row=r, column=1, value=f'=IFERROR(INDEX(Visits!$B$2:$B${last_row},{match_f}),"")')
    ws_r.cell(row=r, column=2, value=f'=IFERROR(INDEX(Visits!$A$2:$A${last_row},{match_f}),"")')
    ws_r.cell(row=r, column=3, value=f'=IFERROR(INDEX(Visits!$C$2:$C${last_row},{match_f}),"")')
    ws_r.cell(row=r, column=4, value=f'=IFERROR(INDEX(Visits!$D$2:$D${last_row},{match_f}),"")')
    ws_r.cell(row=r, column=5, value=f'=IFERROR(INDEX(Visits!$Q$2:$Q${last_row},{match_f}),"")')
    ws_r.cell(row=r, column=6, value=f'=IFERROR(INDEX(Visits!$R$2:$R${last_row},{match_f}),"")')
    ws_r.cell(row=r, column=7, value=f'=IFERROR(INDEX(Visits!$S$2:$S${last_row},{match_f}),"")')
    for cc in range(1,8):
        cell = ws_r.cell(row=r, column=cc)
        set_font(cell, size=9)
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left' if cc>=5 else 'center')
        cell.border = border_all
        if k % 2 == 0:
            fill(cell, LIGHTGREY)

ws_r.column_dimensions['A'].width = 20
ws_r.column_dimensions['B'].width = 20
ws_r.column_dimensions['C'].width = 18
ws_r.column_dimensions['D'].width = 12
ws_r.column_dimensions['E'].width = 45
ws_r.column_dimensions['F'].width = 45
ws_r.column_dimensions['G'].width = 45
ws_r.freeze_panes = "A6"

print("Reflections sheet built")

# reorder sheets: Dashboard, Reflections, Visits, Data, ItemStats, Lists
wb._sheets = [wb["Dashboard"], wb["Reflections"], wb["Visits"], wb["Data"], wb["ItemStats"], wb["Lists"]]
wb.active = 0
ws.sheet_view.tabSelected = True
ws.sheet_view.topLeftCell = "A1"

out_path = "/sessions/affectionate-busy-goodall/mnt/outputs/work/DRF_Centre_Visit_Dashboard.xlsx"
wb.save(out_path)
print("SAVED", out_path)
